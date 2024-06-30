import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import deepcopy
from datetime import datetime,timedelta
import os
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage


# ——————————————————数据部分————————————————————————————————————
source_folder = os.path.join(os.getcwd(), 'PI')
target_folder = os.path.join(os.getcwd(), 'FI')

xlsx_files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx')]
# 读取文件
file_path = os.path.join(source_folder, xlsx_files[0])
df = pd.read_excel(file_path)


# 颜色结束行
null_values = df.iloc[10:, 3].isnull()
color_end_row = null_values.idxmax() if null_values.any() else None

# 尺码结束列
null_value = df.iloc[9, 5:].isnull()
size_end_col = null_value.idxmax() if null_value.any() else None
cols = df.columns
size_col_index = cols.get_loc(size_end_col)
# 尺码
size = df.iloc[9, 4:size_col_index].tolist()

# 有效区间[7行：颜色结束行，所有列]
subset = df.iloc[10:color_end_row, :]

# 工厂
subset.iloc[:, -1] = subset.iloc[:, -1].ffill()
# 款号
sku_only_name = subset.iloc[:, 0].dropna().tolist()
print(sku_only_name)

subset.iloc[:, 0] = subset.iloc[:, 0] .ffill()
# 工厂名字
factory_name = subset.iloc[: , -1].unique().tolist()

# 创建一个空列表来存储每个工厂的过滤数据框
filtered_dfs = []

# 遍历工厂名称列表
for specific_value in factory_name:
    filtered_df = subset[subset.iloc[:, -1]== specific_value]
    filtered_dfs.append(filtered_df)


def download_all_images_from_sheet(filename, save_dir):
    wb = load_workbook(filename)
    # 获取第一个工作表
    sheet = wb.worksheets[0]

    for i, image1 in enumerate(sheet._images, start=1):
        image_anchor = image1.anchor
        image_from = image_anchor._from

        col, colOff, row, rowOff = image_from.col, image_from.colOff, image_from.row, image_from.rowOff
        if col < 1:
            cell_value = str(row)  # 如果是0列，使用行号来命名图片
        else:
            cell_value = sheet.cell(row+1, col).value
        # 获取图像的数据
        img = Image.open(image1.ref).convert("RGB")

        # 将图像数据保存到文件
        save_path = os.path.join(save_dir, f"{cell_value}.png")  # 使用行列信息来保存图片，确保顺序
        img.save(save_path)

        print(f"已下载图像: {save_path}")

    wb.close()

save_dir = 'img'

download_all_images_from_sheet(file_path, save_dir)



#————————————————————————————————————工厂订单部分————————————————————————

#——————————————————————————————————大循环————————————————————————————————————
runtimes = len(filtered_dfs)

name = ['鑫业','众诚','亿多得','邦佐维']

for times in range(runtimes):
     # 文件名
    factory = factory_name[times]
    if factory in name:
        wb = load_workbook(f'{factory}.xlsx')
    else:
        wb = load_workbook('template.xlsx')

    ws = wb.active
    print(f"运行{times}次")
    # 日期
    today = datetime.now()

    # 将当前日期格式化并写入单元格
    ws['P5'] = f'订单日期：{today.strftime("%Y-%m-%d")}'

    # 计算两个月后的日期
    two_months_later = today + timedelta(days=60)

    # 将两个月后的日期格式化并写入单元格
    ws['P6'] = f'交货期：{two_months_later.strftime("%Y-%m-%d")}'


    # 插入尺码
    for i, value in enumerate(size, start=0):
        ws.cell(row=10, column=4 + i, value=value)

    # 文件内容 dataframe
    f1 = filtered_dfs[times]
    df_length = len(f1)

    # 插入指定行数
    if df_length - 1 > 0:
        ws.insert_rows(12, df_length - 1)


    # 拷贝格式
    def copy_cell(copy_from, paste_to_cell):
        # 记录边缘值
        for _copy_row in copy_from:  # 循环每一行
            for _row_cell in _copy_row:  # 循环每一列
                #paste_to_cell.value = _row_cell.value
                paste_to_cell._style = deepcopy(_row_cell._style)  # 复制样式
                paste_to_cell = paste_to_cell.offset(row=0, column=1)  # 右移1格
            paste_to_cell = paste_to_cell.offset(row=1, column=-len(_copy_row))


    for i in range(12, df_length + 11):
        copy_cell(ws['A11':'Q11'], ws[f'A{i}'])
        

    # 插入尺码--一行一行加的
    size_num = f1.iloc[0, 4:size_col_index].tolist()

    for row_index in range(df_length):
        size_num = f1.iloc[row_index, 4:size_col_index].tolist()
        for i, value in enumerate(size_num, start=0):
            ws.cell(row=11 + row_index, column=4 + i, value=value)

    #尺码求和
    size_col_letter = get_column_letter(size_col_index-1)
    end_row = 11 + df_length
    # 用于合计的最后一行
    heji_end_row = 10 + df_length
    # 尺码合计在O列
    for i in range (df_length):
        cell_ref = 11 + i
        ws[f'O{cell_ref}'] = f"=SUM(D{cell_ref}:{size_col_letter}{cell_ref})"

    ws[f'O{end_row}'] = f'=SUM(O11:O{heji_end_row})'

    # 价格求和,价格在P列
    for i in range (df_length):
        cell_ref = 11 + i
        ws[f'P{cell_ref}'] = f"=O{cell_ref}*N{cell_ref}"

    ws[f'P{end_row}'] = f'=SUM(P11:P{heji_end_row})'

    # 定义列索引
    column_indices = {
        '款号': 0,
        '颜色': 3,
        '单价': -2
    }

    # 定义列对应的列号
    column_numbers = {
        '款号': 1,
        '颜色': 13,
        '单价': 14
    }

    for column_name, column_index in column_indices.items():
        values = f1.iloc[:, column_index].tolist()
        for i, value in enumerate(values, start=11):
            ws.cell(row=i, column=column_numbers[column_name], value=value)

    #————————————————————————图片——————————————————————————————

    icoin = ExcelImage('img/0.png')
    ws.add_image(icoin, 'A1')
    new_height = 80  # 设置新高度
    aspect_ratio = icoin.width / icoin.height  # 计算原始宽高比
    new_width = int(new_height * aspect_ratio)  # 根据宽高比计算新宽度
    icoin.width = new_width  # 设置新宽度
    icoin.height = new_height  # 设置新高度


    icoin = ExcelImage('img/None.png')
    ws.add_image(icoin, f'O{heji_end_row + 3}')
    new_height = 120  # 设置新高度
    aspect_ratio = icoin.width / icoin.height  # 计算原始宽高比
    new_width = int(new_height * aspect_ratio)  # 根据宽高比计算新宽度
    icoin.width = new_width  # 设置新宽度
    icoin.height = new_height  # 设置新高度



    img_folder = 'img/'
    processed_files = {}  # 记录已处理过的图片文件名
    for row in range(11, heji_end_row + 1):
        # 获取单元格 A 列的值
        cell_value = ws[f'A{row}'].value
        
        if cell_value is not None and cell_value not in processed_files:
            # 构建图片文件的完整路径
            img_file = img_folder + cell_value + '.png'
            
            # 创建图片对象
            img = ExcelImage(img_file)
            
            # 图片插入范围
            anchor = f'C{row}'
            
            # 设置新高度
            new_height = 100  # 设置新高度
            aspect_ratio = img.width / img.height  # 计算原始宽高比
            new_width = int(new_height * aspect_ratio)  # 根据宽高比计算新宽度
            img.width = new_width  # 设置新宽度
            img.height = new_height  # 设置新高度

            # 插入图片
            ws.add_image(img, anchor)
            
            # 将文件名添加到已处理列表中
            processed_files[cell_value] = True

    # ——————————————————格式部分——————————————————

    type_list = f1.iloc[:, 0].tolist()

    end_row = df_length + 10  # 假设数据从第11行开始

    # 合并单元格
    s = 11
    e = 11
    flag = type_list[0]
    for i in range(1, len(type_list)):
        if type_list[i] != flag:
            flag = type_list[i]
            e = i + 10
            if e >= s:
                ws.merge_cells(f"A{s}:A{e}")
                s = e + 1
        if i == len(type_list) - 1:
            e = end_row
            ws.merge_cells(f"A{s}:A{e}")

    # ————————————————复制合并单元格——————————————————————————————————

    def copy_merged_cells(source_column, source_start_row, source_end_row, target_column):
        merged_cells_in_source_range = [cell for cell in ws.merged_cells.ranges if
                                        cell.min_col == 1 and cell.min_row >= source_start_row and cell.max_row <= source_end_row]

        for cell in merged_cells_in_source_range:
            target_range = f"{target_column}{cell.min_row}:{target_column}{cell.max_row}"
            ws.merge_cells(target_range)

    copy_merged_cells( 'A', 11, end_row + 11, 'B')
    copy_merged_cells( 'A', 11, end_row + 11, 'C')

    # ————————————————————————行高——————————————————————
    merged_ranges = ws.merged_cells.ranges

    for merged_range in merged_ranges:

        start_row = merged_range.min_row
        end_row_in_range = merged_range.max_row
        
        if 11 <= start_row <= end_row and 11 <= end_row_in_range <= end_row:
            # 计算行高
            height = 90 / (end_row_in_range - start_row + 1)
            # 设置行高
            for row_index in range(start_row, end_row_in_range + 1):
                ws.row_dimensions[row_index].height = height

    folder_path = 'FI'
    file_path = os.path.join(folder_path, f'{factory}-GC.xlsx')
    wb.save(file_path)